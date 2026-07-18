"""Trusted XLSX exports for validated Evidence Studio data tables."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from open_notebook.studio.schemas import DataTableDocument

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_INTEGER = re.compile(r"(?:0|[1-9]\d*)$")
_DECIMAL = re.compile(r"(?:0|[1-9]\d*)\.\d+$")


def _typed_value(value: str) -> object:
    """Infer safe Excel types while treating untrusted formulas as text."""
    text = value.strip()
    if text.startswith(_FORMULA_PREFIXES):
        # Excel hides the leading apostrophe, but retains text semantics.
        return f"'{value}"
    if _INTEGER.fullmatch(text) and not (len(text) > 1 and text.startswith("0")):
        return int(text)
    if _DECIMAL.fullmatch(text):
        try:
            return float(Decimal(text))
        except InvalidOperation:
            return value
    try:
        if "T" in text:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        return date.fromisoformat(text)
    except ValueError:
        return value


def _worksheet_title(title: str) -> str:
    sanitized = re.sub(r"[\\\\/*?:\[\]]", " ", title).strip()
    return (sanitized or "Data table")[:31]


def _source_marker_column(columns: list[str]) -> str:
    base = "Source markers"
    if base not in columns:
        return base
    suffix = " (citations)"
    return (base[: 60 - len(suffix)] + suffix)[:60]


def _add_validated_chart(
    workbook: Workbook, numeric_column: int, row_count: int
) -> None:
    if row_count < 1:
        return
    sheet = workbook.active
    chart = BarChart()
    chart.type = "col"
    chart.title = f"{sheet.cell(row=1, column=numeric_column).value} by row"
    chart.y_axis.title = str(sheet.cell(row=1, column=numeric_column).value)
    chart.x_axis.title = "Rows"
    # These references are built from the controlled worksheet bounds, never
    # from model-authored formulas or ranges.
    values = Reference(sheet, min_col=numeric_column, min_row=1, max_row=row_count + 1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=row_count + 1)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "F2")


def export_spreadsheet(document: DataTableDocument, path: Path) -> None:
    """Write an editable, formula-free XLSX from a validated data-table document."""
    if not isinstance(document, DataTableDocument):
        raise TypeError("export_spreadsheet requires DataTableDocument")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("XLSX export path must end in .xlsx")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _worksheet_title(document.title)
    marker_column = _source_marker_column(document.columns)
    headers = [*document.columns, marker_column]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in document.rows:
        values = [_typed_value(row.values[column]) for column in document.columns]
        sheet.append([*values, " ".join(row.citations)])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    for index, header in enumerate(headers, start=1):
        width = max(
            len(header),
            *(len(str(cell.value or "")) for cell in sheet[get_column_letter(index)]),
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            60, max(12, width + 2)
        )

    for column in range(1, len(document.columns) + 1):
        data_cells = [
            sheet.cell(row=row, column=column) for row in range(2, sheet.max_row + 1)
        ]
        if data_cells and all(cell.data_type == "n" for cell in data_cells):
            _add_validated_chart(workbook, column, len(document.rows))
            break

    workbook.properties.title = document.title
    workbook.properties.creator = "Open Notebook Plus"
    workbook.properties.subject = "Evidence Studio editable data export"
    workbook.save(path)


__all__ = ["export_spreadsheet"]
